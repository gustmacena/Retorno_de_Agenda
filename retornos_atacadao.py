import sys
import os
import tkinter as tk
from tkinter import messagebox
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time
import openpyxl
import threading
from datetime import datetime
from PIL import Image, ImageTk
import traceback
import shutil
import tempfile

def limpar_arquivos_temporarios():
    try:
        temp_dir = tempfile.gettempdir()
        for item in os.listdir(temp_dir):
            if item.startswith('_MEI'):
                item_path = os.path.join(temp_dir, item)
                try:
                    if os.path.isfile(item_path):
                        os.unlink(item_path)
                    elif os.path.isdir(item_path):
                        shutil.rmtree(item_path)
                except Exception as e:
                    print(f"Erro ao limpar arquivo temporário {item_path}: {e}")
    except Exception as e:
        print(f"Erro ao limpar arquivos temporários: {e}")

class AutomationRetorno:
    def __init__(self, root):
        self.driver = None
        self.base_path = os.path.join(os.path.expanduser('~'), 'Desktop', 'Retorno Atacadão', 'Base Retorno Atacadao.xlsx')
        self.root = root
        self.setup_ui()
        self.start_time = None
        self.is_running = False

    def setup_ui(self):
        # Labels de status
        self.loading_label = tk.Label(self.root, text="", font=("Helvetica", 9))
        self.loading_label.pack(pady=(10, 0), anchor="w")
        
        self.num_cargas_label = tk.Label(self.root, text="", font=("Helvetica", 9))
        self.num_cargas_label.pack(pady=(10, 0), anchor="w")
        
        self.time_label = tk.Label(self.root, text="", font=("Helvetica", 9))
        self.time_label.pack(pady=(10, 0), anchor="w")
        
        self.status_label = tk.Label(self.root, text="", font=("Helvetica", 9))
        self.status_label.pack(pady=(10, 0), anchor="w")

    def get_driver_path(self):
        try:
            chrome_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'chrome')
            if not os.path.exists(chrome_folder):
                os.makedirs(chrome_folder)
            
            options = webdriver.ChromeOptions()
            options.add_argument(f"--user-data-dir={chrome_folder}")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--disable-gpu")
            options.add_argument("--disable-extensions")
            options.add_argument("--disable-infobars")
            options.add_argument("--disable-notifications")
            options.add_argument("--disable-popup-blocking")
            options.add_argument("--disable-extensions")
            return options
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao configurar driver: {str(e)}")
            raise

    def iniciar_navegador(self):
        try:
            options = self.get_driver_path()
            options.add_argument("--start-maximized")
            
            chrome_executable = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'chrome', 'chrome.exe')
            if os.path.exists(chrome_executable):
                options.binary_location = chrome_executable
            
            service = ChromeService()
            self.driver = webdriver.Chrome(service=service, options=options)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao iniciar navegador: {str(e)}")
            raise

    def login(self):
        if not self.driver:
            self.iniciar_navegador()

        print("Realizando login...")
        self.driver.get('https://atacadao.hodiebooking.com.br/cargas')
        WebDriverWait(self.driver, 40).until(EC.url_contains('https://atacadao.hodiebooking.com.br/cargas'))
        print("Login realizado com sucesso!")

        self.realizar_proximos_passos()

    def realizar_proximos_passos(self):
        try:
            self.is_running = True
            self.start_time = datetime.now()
            self.update_time_label()
            self.loading_label.config(text="Iniciando automação...")
            self.status_label.config(text="Status: Em execução")

            numeros_carga = self.ler_numeros_carga()
            if not numeros_carga:
                messagebox.showwarning("Aviso", "Nenhum número de carga encontrado na planilha!")
                return

            self.num_cargas_label.config(text=f"Número de cargas: {len(numeros_carga)}")

            for numero_carga in numeros_carga:
                if not self.is_running:
                    break
                    
                self.status_label.config(text=f"Status: Processando carga {numero_carga}")

                try:
                    self.driver.get('https://atacadao.hodiebooking.com.br/cargas')
                    WebDriverWait(self.driver, 20).until(EC.url_contains('https://atacadao.hodiebooking.com.br/cargas'))
                    time.sleep(2)
                    print("Retornou para a URL de pesquisa.")

                    element = WebDriverWait(self.driver, 20).until(
                        EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[1]/div[1]/div[1]/div/a/div'))
                    )
                    element.click()
                    print("Clicou no campo especificado com sucesso.")

                    time.sleep(2)
                    print("Esperou 2 segundos após o último passo.")

                    if numero_carga:
                        campo_codigo = WebDriverWait(self.driver, 20).until(
                            EC.presence_of_element_located((By.XPATH, '//*[@id="codigo"]'))
                        )
                        campo_codigo.clear()
                        campo_codigo.send_keys(numero_carga)
                        print(f"Inseriu o número da carga '{numero_carga}' no campo especificado.")

                        data_criacao = WebDriverWait(self.driver, 20).until(
                            EC.element_to_be_clickable((By.XPATH, '//*[@id="filtros-collapse"]/div[1]/div/div[5]/div/span[2]'))
                        )
                        data_criacao.click()
                        print("Clicou no botão para limpar a data de criação.")

                        time.sleep(2)
                        print("Esperou 2 segundos após o último passo.")

                        btn_aplicar_filtro = WebDriverWait(self.driver, 20).until(
                            EC.element_to_be_clickable((By.XPATH, '//*[@id="enviarFiltros"]'))
                        )
                        btn_aplicar_filtro.click()
                        print("Clicou no botão 'Aplicar Filtro'.")

                        time.sleep(2)
                        print("Esperou 2 segundos após clicar em 'Aplicar Filtro'.")

                        try:
                            btn_exibir = WebDriverWait(self.driver, 20).until(
                                EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[1]/div[1]/div[3]/div/div/div/div[1]/table/tbody/tr/td[8]/div/a'))
                            )
                            btn_exibir.click()
                            print("Clicou no botão 'Exibir'.")

                            time.sleep(2)
                            print("Esperou 2 segundos após clicar em 'Exibir'.")

                            try:
                                # Coletar nova informação: Data desejada de agendamento
                                data_desejada_element = WebDriverWait(self.driver, 20).until(
                                    EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div[1]/div[4]/form/div[2]/div[1]/div[2]/div/div[2]/div/div[1]/div[2]'))
                                )
                                data_desejada = data_desejada_element.text
                                print(f"Data desejada de agendamento encontrada: {data_desejada}")

                                # Coletar nova informação: Quando foi solicitado?
                                solicitado_element = WebDriverWait(self.driver, 20).until(
                                    EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div[1]/div[4]/form/div[2]/div[1]/div[1]/div/div[2]/div[3]/div[6]/div/div'))
                                )
                                solicitado = solicitado_element.text
                                print(f"Quando foi solicitado? encontrado: {solicitado}")

                                # Clicar no botão 'Acompanhe Agendamento'
                                status_element = WebDriverWait(self.driver, 20).until(
                                    EC.presence_of_element_located((By.XPATH, '//*[@id="cargasForm"]/div[2]/div[1]/div[1]/div/div[1]/div[2]'))
                                )
                                status = status_element.text
                                print(f"Status encontrado: {status}")

                                if "Aguardando aprovação de agendamento" in status:
                                    self.escrever_dados_na_planilha(numero_carga, "", "Aguardando aprovação de agendamento", data_desejada, solicitado, "")
                                elif any(char.isdigit() for char in status):
                                    btn_acompanhe_agendamento = WebDriverWait(self.driver, 20).until(
                                        EC.element_to_be_clickable((By.XPATH, '//*[@id="cargasForm"]/div[2]/div[1]/div[1]/div/div[1]/div[4]/a/i'))
                                    )
                                    btn_acompanhe_agendamento.click()
                                    print("Clicou no botão 'Acompanhe Agendamento'.")

                                    time.sleep(2)
                                    print("Esperou 2 segundos após clicar em 'Acompanhe Agendamento'.")

                                    try:
                                        protocolo_element = WebDriverWait(self.driver, 20).until(
                                            EC.presence_of_element_located((By.XPATH, '//*[@id="agendamentoForm"]/div[2]/div[2]/div[1]/div/div[2]/div[1]/div/div[1]/span'))
                                        )
                                        protocolo = protocolo_element.text
                                        print(f"Protocolo encontrado: {protocolo}")

                                        data_entrega_element = WebDriverWait(self.driver, 20).until(
                                            EC.presence_of_element_located((By.XPATH, '//*[@id="agendamentoForm"]/div[2]/div[2]/div[1]/div/div[2]/div[1]/div/div[2]/span'))
                                        )
                                        data_entrega = data_entrega_element.text.split(' / ')[0]  # Extrai apenas a data
                                        print(f"Data de Entrega Aprovada encontrada: {data_entrega}")

                                        # Coletar nova informação: Solicitante do agendamento
                                        solicitante_element = WebDriverWait(self.driver, 20).until(
                                            EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div[1]/div[2]/form/div[2]/div[1]/div[2]/div/div[1]/div'))
                                        )
                                        solicitante = solicitante_element.text
                                        print(f"Solicitante do agendamento encontrado: {solicitante}")

                                        self.escrever_dados_na_planilha(numero_carga, protocolo, data_entrega, data_desejada, solicitado, solicitante)
                                    except Exception as e:
                                        print(f"Erro ao processar dados de agendamento: {e}")
                                else:
                                    print("Status não é 'Aguardando aprovação de agendamento' ou não contém data aprovada. Pulando para a próxima carga.")
                            except Exception:
                                print("Status não encontrado. Registrando na planilha...")
                                self.escrever_dados_na_planilha(numero_carga, "", "Status não encontrado", data_desejada, solicitado, "")
                        except Exception as e:
                            print(f"Erro ao clicar em 'Exibir': {e}")
                except Exception as e:
                    self.status_label.config(text=f"Status: Erro ao processar carga {numero_carga}")
                    continue

            self.loading_label.config(text="Automação Finalizada")
            self.status_label.config(text="Status: Finalizado")
            self.update_time_label()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro na automação: {str(e)}")
        finally:
            self.is_running = False

    def ler_numeros_carga(self):
        try:
            workbook = openpyxl.load_workbook(self.base_path)
            sheet = workbook.active
            numeros_carga = []
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=4, column=col).value == "Número de Carga":
                    for row in range(5, sheet.max_row + 1):
                        numero_carga = sheet.cell(row=row, column=col).value
                        if numero_carga:
                            numeros_carga.append(numero_carga)
                    break
            return numeros_carga
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ler a planilha: {str(e)}")
            return []

    def escrever_dados_na_planilha(self, numero_carga, protocolo, data_ou_status, data_desejada, solicitado, solicitante):
        try:
            workbook = openpyxl.load_workbook(self.base_path)
            sheet = workbook.active
            for row in range(5, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    if sheet.cell(row=row, column=col).value == numero_carga:
                        sheet.cell(row=row, column=col + 1).value = protocolo
                        sheet.cell(row=row, column=col + 2).value = data_ou_status  # Coluna "Data de agenda confirmada"
                        sheet.cell(row=row, column=col + 3).value = data_desejada
                        sheet.cell(row=row, column=col + 4).value = solicitado
                        sheet.cell(row=row, column=col + 5).value = solicitante
                        break
            workbook.save(self.base_path)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao escrever na planilha: {str(e)}")

    def fechar_navegador(self):
        if self.driver:
            self.driver.quit()

    def update_time_label(self):
        if self.start_time:
            now = datetime.now()
            elapsed_time = now - self.start_time
            elapsed_seconds = int(elapsed_time.total_seconds())
            minutes = elapsed_seconds // 60
            seconds = elapsed_seconds % 60
            self.time_label.config(text=f"Tempo de Execução: {minutes:02}:{seconds:02}")

    def stop_automation(self):
        self.is_running = False
        self.status_label.config(text="Status: Parando...")
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass
        self.status_label.config(text="Status: Parado")

def run_automation(automation_instance):
    automation_instance.login()
    automation_instance.fechar_navegador()

def start_automation():
    global automation_instance
    if automation_instance and automation_instance.is_running:
        messagebox.showwarning("Aviso", "Automação já está em execução!")
        return
        
    automation_instance = AutomationRetorno(root)
    threading.Thread(target=run_automation, args=(automation_instance,)).start()

def stop_automation():
    global automation_instance
    if automation_instance:
        automation_instance.stop_automation()

# Interface Gráfica
root = tk.Tk()
root.title("Automação Atacadão")
root.resizable(width=False, height=False)

# Configurações de cores
bauducco_yellow = "#FEB81C"
bauducco_red = "#FF0000"
bauducco_verde = "#1FB757"
bauducco_black = "#000000"

root.configure(bg=bauducco_yellow)

# Carregar imagens
try:
    left_logo_img = Image.open(os.path.join('imgs', 'Logo_Bauducco.png'))
    left_logo_img = left_logo_img.resize((120, 95), Image.LANCZOS)
    left_logo_photo = ImageTk.PhotoImage(left_logo_img)

    right_logo_img = Image.open(os.path.join('imgs', 'Logo_Atacadao.png'))
    right_logo_img = right_logo_img.resize((120, 95), Image.LANCZOS)
    right_logo_photo = ImageTk.PhotoImage(right_logo_img)
except Exception as e:
    messagebox.showerror("Erro", "Erro ao carregar imagens do programa")

# Logo e Títulos 
top_frame = tk.Frame(root, bg=bauducco_yellow)
top_frame.pack(pady=10, fill="x")

left_logo = tk.Label(top_frame, image=left_logo_photo, bg=bauducco_yellow)
left_logo.pack(side="left", padx=20)

right_logo = tk.Label(top_frame, image=right_logo_photo, bg=bauducco_yellow)
right_logo.pack(side="right", padx=20)

title_label = tk.Label(root, text="Retorno de agenda", font=("Helvetica", 15, "bold"), bg=bauducco_yellow, fg=bauducco_black)
title_label.pack(pady=(10, 20))

dev_label = tk.Label(root, text="Desenvolvedor: Gustavo Macena", font=("Helvetica", 10), bg=bauducco_yellow, fg=bauducco_black)
dev_label.pack(pady=(1, 1), anchor="w")

dev_label = tk.Label(root, text="Equipe: Projetos de melhoria continua", font=("Helvetica", 10), bg=bauducco_yellow, fg=bauducco_black)
dev_label.pack(pady=(1,10), anchor="w")

# Frame para botões
button_frame = tk.Frame(root, bg=bauducco_yellow)
button_frame.pack(pady=20)

# Botões
start_button = tk.Button(button_frame, text="Iniciar", font=("Helvetica", 14), bg=bauducco_verde, fg=bauducco_black, command=start_automation)
start_button.pack(side=tk.LEFT, padx=10)

stop_button = tk.Button(button_frame, text="Parar", font=("Helvetica", 14), bg=bauducco_red, fg=bauducco_black, command=stop_automation)
stop_button.pack(side=tk.LEFT, padx=10)

# Label de Loading
automation_instance = None

# Tratamento de fechamento da janela
def on_closing():
    if automation_instance and automation_instance.is_running:
        if messagebox.askokcancel("Confirmar", "A automação está em execução. Deseja realmente sair?"):
            automation_instance.stop_automation()
            limpar_arquivos_temporarios()
            root.destroy()
    else:
        limpar_arquivos_temporarios()
        root.destroy()

root.protocol("WM_DELETE_WINDOW", on_closing)

# Main loop da interface
root.mainloop()